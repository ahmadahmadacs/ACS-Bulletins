"""
ACS — Family Matcher  v2
Lie les élèves Schoolify au fichier notes via matching multi-critères :
  1. Exact
  2. Normalisation arabe + prénom + dernier mot famille
  3. Prénom normalisé + groupe EB unique
  4. Famille sans espaces (ابو عمشة → ابوعمشة)
  5. Prénom 2 mots collés (عبد الرحيم → عبدالرحيم)
  6. Famille 2 mots collés (عبد القادر → عبدالقادر)
  7. Fuzzy Levenshtein ≤ 1 (fautes de frappe)
  8. Correspondances manuelles fixes
"""

import re
from openpyxl import load_workbook

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
GRADE_TO_GROUPE = {
    'Grade 1': 'EB1-2', 'Grade 2': 'EB1-2',
    'EB3': 'EB3-6', 'EB4': 'EB3-6', 'EB5': 'EB3-6', 'EB6': 'EB3-6',
    'EB7': 'EB7', 'EB8': 'EB8', 'EB9': 'EB9',
}

SHEET_TO_GROUPE = {
    'S1 Notes EB1-2': 'EB1-2', 'S1 Notes EB3-6': 'EB3-6',
    'S1 Notes EB7':   'EB7',   'S1 Notes EB8':   'EB8',
    'S1 Notes EB9':   'EB9',
}

# Correspondances manuelles : sc_name → nom exact dans fichier notes
MANUAL_OVERRIDES = {
    'زيد وليد عبدالله':     'زيد عبدالله',
    'جوري أحمد خضر الادرع': 'جوري خضر',
    'يحيى احمد خضر':        'يحيى خضر',
    'عبد الرحيم احمد خضر':  'عبدالرحيم خضر',
    'تالين خضر عبدالله':    'تالين عبدلله',
    # سدره ايهاب دياب → absente du fichier notes
}

# ─────────────────────────────────────────────
# NORMALISATION
# ─────────────────────────────────────────────
def normalize_arabic(text: str) -> str:
    text = re.sub(r'[\u064B-\u065F\u0670]', '', text)
    text = re.sub(r'[أإآ]', 'ا', text)
    text = text.replace('ة', 'ه').replace('ى', 'ي')
    return re.sub(r'\s+', ' ', text).strip()

def collapse_spaces(text: str) -> str:
    return text.replace(' ', '')

def levenshtein(a: str, b: str) -> int:
    if abs(len(a) - len(b)) > 2:
        return 99
    la, lb = len(a), len(b)
    dp = list(range(lb + 1))
    for i in range(1, la + 1):
        prev = dp[:]
        dp[0] = i
        for j in range(1, lb + 1):
            cost = 0 if a[i-1] == b[j-1] else 1
            dp[j] = min(dp[j] + 1, dp[j-1] + 1, prev[j-1] + cost)
    return dp[lb]

# ─────────────────────────────────────────────
# CHARGEMENT
# ─────────────────────────────────────────────
def load_notes_students(notes_path) -> dict:
    wb = load_workbook(notes_path, read_only=True, data_only=True)
    students = {}
    for sheet, groupe in SHEET_TO_GROUPE.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue
            nom, classe = row[0], row[1]
            if nom and isinstance(nom, str) and 'معدل' not in nom and nom != 'Nom':
                students[nom.strip()] = {
                    'groupe': groupe,
                    'classe': str(classe).strip() if classe else '',
                }
    wb.close()
    return students

def load_schoolify_students(schoolify_path) -> list:
    wb = load_workbook(schoolify_path, read_only=True, data_only=True)
    ws = wb['Students']
    students = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        student_id, name, prenom, pere, famille, surname, parent_id, grade, cycle, job = (list(row) + [None]*10)[:10]
        if not name or not parent_id:
            continue
        if cycle and 'Maternelle' in str(cycle):
            continue
        students.append({
            'name':       str(name).strip(),
            'prenom':     str(prenom).strip()     if prenom     else '',
            'pere':       str(pere).strip()       if pere       else '',
            'famille':    str(famille).strip()    if famille    else '',
            'parent_id':  str(parent_id).strip(),
            'grade':      str(grade).strip()      if grade      else '',
            'job':        str(job).strip()        if job        else '',
            'student_id': str(student_id).strip() if student_id else '',
        })
    wb.close()
    return students

# ─────────────────────────────────────────────
# INDEX
# ─────────────────────────────────────────────
def build_indices(notes_students: dict) -> dict:
    idx = {
        'exact':      set(notes_students.keys()),
        'pf':         {},
        'pg':         {},
        'pf_nospace': {},
        'p2f':        {},
        'pf2':        {},
        'norm_full':  {},
    }
    for nom, info in notes_students.items():
        g = info['groupe']
        parts = normalize_arabic(nom).split()
        if not parts:
            continue
        p, f = parts[0], parts[-1]
        f_nospace = collapse_spaces(normalize_arabic(nom.split()[-1]))
        full_norm = collapse_spaces(normalize_arabic(nom))

        idx['pf'].setdefault((p, f), []).append(nom)
        idx['pg'].setdefault((p, g), []).append(nom)
        idx['pf_nospace'].setdefault((p, f_nospace), []).append(nom)
        idx['norm_full'][full_norm] = nom

        if len(parts) >= 3:
            p2 = parts[0] + parts[1]
            idx['p2f'].setdefault((p2, f), []).append(nom)

        if len(nom.split()) >= 4:
            raw = nom.split()
            fam2 = normalize_arabic(''.join(raw[-2:]))
            idx['pf2'].setdefault((p, fam2), []).append(nom)

    return idx

# ─────────────────────────────────────────────
# MATCHING
# ─────────────────────────────────────────────
def match_one(sc: dict, notes_students: dict, idx: dict) -> tuple:
    name   = sc['name']
    grade  = sc['grade']
    groupe = GRADE_TO_GROUPE.get(grade, '')

    n_name = normalize_arabic(name)
    parts  = n_name.split()
    if not parts:
        return None, 'empty_name'

    p = parts[0]
    f = parts[-1]
    f_col     = normalize_arabic(sc['famille']) if sc['famille'] else f
    f_nospace = collapse_spaces(f_col)

    def filter_groupe(cands):
        if not groupe:
            return cands
        return [c for c in cands if notes_students[c]['groupe'] == groupe]

    # Passe 1 — exact
    if name in idx['exact']:
        return name, 'exact'

    # Passe 2 — norm prénom + famille
    for fam in set([f, f_col]):
        cands = idx['pf'].get((p, fam), [])
        if len(cands) == 1:
            return cands[0], 'norm_pf'
        fg = filter_groupe(cands)
        if len(fg) == 1:
            return fg[0], 'norm_pf+groupe'

    # Passe 3 — norm prénom unique dans groupe
    if groupe:
        cands = idx['pg'].get((p, groupe), [])
        if len(cands) == 1:
            return cands[0], 'prenom_unique_groupe'

    # Passe 4 — famille sans espaces
    cands = idx['pf_nospace'].get((p, f_nospace), [])
    if len(cands) == 1:
        return cands[0], 'nospace_famille'
    fg = filter_groupe(cands)
    if len(fg) == 1:
        return fg[0], 'nospace_famille+groupe'

    # Passe 5 — prénom 2 mots collés
    if len(parts) >= 3:
        p2 = parts[0] + parts[1]
        cands = idx['p2f'].get((p2, f), [])
        if len(cands) == 1:
            return cands[0], 'prenom2mots'
        fg = filter_groupe(cands)
        if len(fg) == 1:
            return fg[0], 'prenom2mots+groupe'

    # Passe 6 — famille 2 mots collés
    if len(name.split()) >= 4:
        fam2 = normalize_arabic(''.join(name.split()[-2:]))
        cands = idx['pf2'].get((p, fam2), [])
        if len(cands) == 1:
            return cands[0], 'famille2mots'
        fg = filter_groupe(cands)
        if len(fg) == 1:
            return fg[0], 'famille2mots+groupe'

    # Passe 7 — Levenshtein ≤ 1
    full_norm = collapse_spaces(n_name)
    best_match, best_dist = None, 2
    for full_n, nom_notes in idx['norm_full'].items():
        if groupe and notes_students[nom_notes]['groupe'] != groupe:
            continue
        d = levenshtein(full_norm, full_n)
        if d < best_dist:
            best_dist, best_match = d, nom_notes
    if best_match and best_dist <= 1:
        return best_match, f'fuzzy_lev{best_dist}'

    return None, 'UNMATCHED'

# ─────────────────────────────────────────────
# POINT D'ENTRÉE
# ─────────────────────────────────────────────
def match_families(notes_path, schoolify_path) -> tuple:
    """
    Retourne :
        family_index   : { parent_id → [nom_notes, ...] }
        unmatched      : liste de dicts pour les élèves non matchés
        name_mapping   : { sc_name → notes_name }
        notes_students : { nom_notes → { groupe, classe } }
    """
    notes_students = load_notes_students(notes_path)
    sc_students    = load_schoolify_students(schoolify_path)
    idx            = build_indices(notes_students)

    name_mapping = {}
    unmatched    = []

    for sc in sc_students:
        # Override manuel prioritaire
        if sc['name'] in MANUAL_OVERRIDES:
            notes_name = MANUAL_OVERRIDES[sc['name']]
            if notes_name in notes_students:
                name_mapping[sc['name']] = notes_name
                continue
        result, method = match_one(sc, notes_students, idx)
        if result:
            name_mapping[sc['name']] = result
        else:
            unmatched.append({
                'schoolify_name': sc['name'],
                'grade':          sc['grade'],
                'parent_id':      sc['parent_id'],
                'job':            sc['job'],
                'method':         method,
            })

    # Index famille → noms notes
    family_index = {}
    for sc in sc_students:
        nm = name_mapping.get(sc['name'])
        if nm:
            pid = sc['parent_id']
            family_index.setdefault(pid, [])
            if nm not in family_index[pid]:
                family_index[pid].append(nm)

    return family_index, unmatched, name_mapping, notes_students


# ─────────────────────────────────────────────
# LECTURE DIRECTE DEPUIS FICHIER NOTES ENRICHI
# ─────────────────────────────────────────────
def read_family_index_from_notes(notes_path) -> tuple:
    """
    Lit les ID_Parent directement depuis le fichier notes enrichi.
    Utilise la colonne 'ID_Parent' ajoutée par inject_parent_ids.py.

    Retourne :
        family_index   : { parent_id → [nom_notes, ...] }
        notes_students : { nom_notes → { groupe, classe } }
        has_parent_col : bool — True si la colonne ID_Parent est présente
    """
    wb = load_workbook(notes_path, read_only=True, data_only=True)
    notes_students = {}
    family_index   = {}
    has_parent_col = False

    for sheet, groupe in SHEET_TO_GROUPE.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]

        # Trouver la colonne ID_Parent dans les 3 premières lignes
        id_parent_col = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 3:
                break
            for j, v in enumerate(row):
                if v == 'ID_Parent':
                    id_parent_col = j  # 0-based
                    has_parent_col = True
                    break
            if id_parent_col is not None:
                break

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 1:
                continue
            nom = row[0]
            if not nom or not isinstance(nom, str):
                continue
            nom = nom.strip()
            if 'معدل' in nom or nom in ('Nom',):
                continue

            classe = row[1] if len(row) > 1 else ''
            notes_students[nom] = {
                'groupe': groupe,
                'classe': str(classe).strip() if classe else '',
            }

            if id_parent_col is not None and id_parent_col < len(row):
                pid = row[id_parent_col]
                if pid and isinstance(pid, str):
                    pid = pid.strip()
                    family_index.setdefault(pid, [])
                    if nom not in family_index[pid]:
                        family_index[pid].append(nom)

    wb.close()
    return family_index, notes_students, has_parent_col
