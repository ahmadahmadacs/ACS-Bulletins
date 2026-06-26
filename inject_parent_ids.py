"""
ACS — inject_parent_ids.py
Script à lancer UNE SEULE FOIS pour enrichir le fichier notes
avec les ID parents depuis Schoolify.

Usage :
    python inject_parent_ids.py \
        --notes  Bulletin_notes_25-26_S3_new_design.xlsx \
        --schoolify Schoolify_Data_V1.xlsx \
        --out    Bulletin_notes_25-26_avec_parents.xlsx

Le script ajoute une colonne "ID_Parent" après النتيجة
dans TOUTES les feuilles "SX Notes EBX".
Aucune donnée existante n'est modifiée.
"""

import argparse
import shutil
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ── Normalisation arabe (même logique que family_matcher) ──────────────
def normalize_arabic(text: str) -> str:
    text = re.sub(r'[\u064B-\u065F\u0670]', '', text)
    text = re.sub(r'[أإآ]', 'ا', text)
    text = text.replace('ة', 'ه').replace('ى', 'ي')
    return re.sub(r'\s+', ' ', text).strip()

def norm_key(name: str) -> str:
    parts = normalize_arabic(name).split()
    return (parts[0] + ' ' + parts[-1]) if len(parts) >= 2 else name

def collapse_spaces(text: str) -> str:
    return text.replace(' ', '')

def levenshtein(a: str, b: str) -> int:
    if abs(len(a) - len(b)) > 2:
        return 99
    la, lb = len(a), len(b)
    dp = list(range(lb + 1))
    for i in range(1, la + 1):
        prev = dp[:]
        dp[0] = i
        for j in range(1, lb + 1):
            cost = 0 if a[i-1] == b[j-1] else 1
            dp[j] = min(dp[j] + 1, dp[j-1] + 1, prev[j-1] + cost)
    return dp[lb]

# ── Overrides manuels (cas non résolvables automatiquement) ────────────
MANUAL_OVERRIDES = {
    'زيد وليد عبدالله':     'زيد عبدالله',
    'جوري أحمد خضر الادرع': 'جوري خضر',
    'يحيى احمد خضر':        'يحيى خضر',
    'عبد الرحيم احمد خضر':  'عبدالرحيم خضر',
    'تالين خضر عبدالله':    'تالين عبدلله',
}

GRADE_TO_GROUPE = {
    'Grade 1': 'EB1-2', 'Grade 2': 'EB1-2',
    'EB3': 'EB3-6', 'EB4': 'EB3-6', 'EB5': 'EB3-6', 'EB6': 'EB3-6',
    'EB7': 'EB7', 'EB8': 'EB8', 'EB9': 'EB9',
}

NOTE_SHEETS = [
    'S1 Notes EB1-2', 'S2 Notes EB1-2', 'S3 Notes EB1-2',
    'S1 Notes EB3-6', 'S2 Notes EB3-6', 'S3 Notes EB3-6',
    'S1 Notes EB7',   'S2 Notes EB7',   'S3 Notes EB7',
    'S1 Notes EB8',   'S2 Notes EB8',   'S3 Notes EB8',
    'S1 Notes EB9',   'S2 Notes EB9',   'S3 Notes EB9',
]

SHEET_TO_GROUPE = {
    'S1 Notes EB1-2': 'EB1-2', 'S2 Notes EB1-2': 'EB1-2', 'S3 Notes EB1-2': 'EB1-2',
    'S1 Notes EB3-6': 'EB3-6', 'S2 Notes EB3-6': 'EB3-6', 'S3 Notes EB3-6': 'EB3-6',
    'S1 Notes EB7':   'EB7',   'S2 Notes EB7':   'EB7',   'S3 Notes EB7':   'EB7',
    'S1 Notes EB8':   'EB8',   'S2 Notes EB8':   'EB8',   'S3 Notes EB8':   'EB8',
    'S1 Notes EB9':   'EB9',   'S2 Notes EB9':   'EB9',   'S3 Notes EB9':   'EB9',
}


# ── Chargement Schoolify → mapping nom_notes → parent_id ──────────────
def build_parent_mapping(schoolify_path: Path, notes_names: set) -> dict:
    """Retourne { nom_notes → parent_id }"""
    wb = load_workbook(schoolify_path, read_only=True, data_only=True)
    ws = wb['Students']

    sc_students = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        row = list(row) + [None] * 10
        student_id, name, prenom, pere, famille, surname, parent_id, grade, cycle, job = row[:10]
        if not name or not parent_id:
            continue
        if cycle and 'Maternelle' in str(cycle):
            continue
        sc_students.append({
            'name':      str(name).strip(),
            'famille':   str(famille).strip() if famille else '',
            'parent_id': str(parent_id).strip(),
            'grade':     str(grade).strip() if grade else '',
        })
    wb.close()

    # Construire index pour matching
    def norm_p(n): return normalize_arabic(n).split()[0]
    def norm_f(n): return normalize_arabic(n).split()[-1]

    idx_pf  = {}
    idx_pg  = {}
    idx_full= {}
    for nom in notes_names:
        p, f  = norm_p(nom), norm_f(nom)
        parts = normalize_arabic(nom).split()
        g     = _guess_groupe_from_notes(nom)
        idx_pf.setdefault((p, f), []).append(nom)
        idx_pg.setdefault((p, g), []).append(nom)
        idx_full[collapse_spaces(normalize_arabic(nom))] = nom

    def find_notes_name(sc):
        name   = sc['name']
        grade  = sc['grade']
        groupe = GRADE_TO_GROUPE.get(grade, '')
        f_col  = normalize_arabic(sc['famille']) if sc['famille'] else ''
        n      = normalize_arabic(name)
        parts  = n.split()
        if not parts: return None
        p, f   = parts[0], parts[-1]

        if name in notes_names: return name

        for fam in set([f, f_col]):
            cands = idx_pf.get((p, fam), [])
            if len(cands) == 1: return cands[0]
            if groupe:
                fg = [c for c in cands if _guess_groupe_from_notes(c) == groupe]
                if len(fg) == 1: return fg[0]

        if groupe:
            cands = idx_pg.get((p, groupe), [])
            if len(cands) == 1: return cands[0]

        f_ns = collapse_spaces(f_col or f)
        cands = [nn for nn, nf in [(nn, norm_f(nn)) for nn in notes_names]
                 if norm_p(nn) == p and collapse_spaces(nf) == f_ns]
        if len(cands) == 1: return cands[0]

        if len(parts) >= 3:
            p2 = parts[0] + parts[1]
            cands = [nn for nn in notes_names if collapse_spaces(normalize_arabic(nn)).startswith(p2) and norm_f(nn) == f]
            if len(cands) == 1: return cands[0]

        full = collapse_spaces(n)
        best, bdist = None, 2
        for fn, nn in idx_full.items():
            if groupe and _guess_groupe_from_notes(nn) != groupe: continue
            d = levenshtein(full, fn)
            if d < bdist: bdist, best = d, nn
        if best and bdist <= 1: return best

        return None

    # Détecter les doublons de noms dans Schoolify
    from collections import Counter
    nom_count = Counter(sc['name'] for sc in sc_students)
    doublons = {n for n, c in nom_count.items() if c > 1}
    if doublons:
        print(f"  ⚠ Noms en doublon détectés : {doublons}")

    # mapping (nom_notes, classe_arabe) → parent_id  pour les doublons
    # mapping nom_notes → parent_id  pour les cas simples
    mapping       = {}    # nom_notes → parent_id  (cas simples)
    mapping_exact = {}    # (nom_notes, classe_arabe) → parent_id  (doublons)
    unmatched = []

    for sc in sc_students:
        # Override manuel
        override_name = MANUAL_OVERRIDES.get(sc['name'])
        if override_name and override_name in notes_names:
            if sc['name'] in doublons:
                classes = GRADE_TO_CLASSE.get(sc['grade'], [])
                for cl in classes:
                    mapping_exact[(override_name, cl)] = sc['parent_id']
            else:
                mapping[override_name] = sc['parent_id']
            continue

        notes_name = find_notes_name(sc)
        if notes_name:
            if sc['name'] in doublons:
                classes = GRADE_TO_CLASSE.get(sc['grade'], [])
                for cl in classes:
                    mapping_exact[(notes_name, cl)] = sc['parent_id']
            else:
                mapping[notes_name] = sc['parent_id']
        else:
            unmatched.append(sc['name'])

    return mapping, mapping_exact, unmatched


# Heuristique groupe depuis le nom (on n'a pas le groupe ici — on le déduit du contexte)
_groupe_cache = {}

def _guess_groupe_from_notes(nom: str) -> str:
    """Retourne le groupe d'un élève depuis le cache construit lors du scan."""
    return _groupe_cache.get(nom, '')


# ── Injection dans le fichier notes ───────────────────────────────────
def inject_parent_ids(notes_path: Path, schoolify_path: Path, out_path: Path):
    print(f"📂 Lecture du fichier notes : {notes_path.name}")
    shutil.copy(notes_path, out_path)
    wb = load_workbook(out_path)

    # Passe 1 — collecter tous les noms + groupe depuis les feuilles S1
    all_notes_names = set()
    for sheet, groupe in SHEET_TO_GROUPE.items():
        if sheet not in wb.sheetnames or not sheet.startswith('S1'):
            continue
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 1: continue
            nom = row[0]
            if nom and isinstance(nom, str) and 'معدل' not in nom and nom not in ('Nom',):
                nom = nom.strip()
                all_notes_names.add(nom)
                _groupe_cache[nom] = groupe

    print(f"  → {len(all_notes_names)} élèves trouvés dans le fichier notes")

    # Passe 2 — construire le mapping nom → parent_id
    print(f"📂 Lecture Schoolify : {schoolify_path.name}")
    parent_map, parent_map_exact, unmatched = build_parent_mapping(schoolify_path, all_notes_names)
    print(f"  → {len(parent_map) + len(parent_map_exact)} élèves matchés · {len(unmatched)} non matchés")
    if unmatched:
        print(f"  ⚠ Non matchés : {unmatched}")

    # Passe 3 — injecter dans chaque feuille
    HEADER_STYLE = Font(name='Calibri', bold=True, color='FFFFFFFF')
    HEADER_FILL  = PatternFill('solid', fgColor='1A3A6B')
    CELL_FONT    = Font(name='Calibri', size=10)
    CELL_ALIGN   = Alignment(horizontal='center', vertical='center')

    injected_total = 0

    for sheet in wb.sheetnames:
        if sheet not in SHEET_TO_GROUPE:
            continue
        ws = wb[sheet]

        # Trouver la colonne النتيجة (header) → ID_Parent sera juste après
        result_col = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 3: break
            for j, v in enumerate(row):
                if v == 'النتيجة':
                    result_col = j + 1  # 1-based
                    break
            if result_col:
                break

        if result_col is None:
            # Fallback : utiliser max_column + 1 (mais plafonné à 20)
            result_col = min(ws.max_column or 15, 20)
        
        id_col = result_col + 1  # colonne ID_Parent

        # Écrire le header "ID_Parent" sur la même ligne que النتيجة
        for i, row in enumerate(ws.iter_rows()):
            if i > 3: break
            for cell in row:
                if cell.value == 'النتيجة':
                    header_cell = ws.cell(row=cell.row, column=id_col)
                    header_cell.value = 'ID_Parent'
                    header_cell.font      = HEADER_STYLE
                    header_cell.fill      = HEADER_FILL
                    header_cell.alignment = CELL_ALIGN
                    break

        # Remplir ID_Parent pour chaque élève
        injected = 0
        for row in ws.iter_rows():
            nom_cell = row[0]
            nom = nom_cell.value
            if not nom or not isinstance(nom, str):
                continue
            nom = nom.strip()
            if 'معدل' in nom or nom in ('Nom',):
                continue

            # Lire la classe (col B) pour résoudre les doublons
            classe_cell = row[1] if len(row) > 1 else None
            classe = str(classe_cell.value).strip() if classe_cell and classe_cell.value else ''

            # Chercher d'abord dans l'index exact (nom+classe) pour les doublons
            pid = parent_map_exact.get((nom, classe)) or parent_map.get(nom)

            if pid:
                id_cell = ws.cell(row=nom_cell.row, column=id_col)
                id_cell.value     = pid
                id_cell.font      = CELL_FONT
                id_cell.alignment = CELL_ALIGN
                injected += 1

        injected_total += injected
        print(f"  ✔ {sheet} — {injected} IDs injectés (col {id_col})")

    wb.save(out_path)
    wb.close()
    print(f"\n✅ Fichier enrichi sauvegardé : {out_path.name}")
    print(f"   {injected_total} cellules ID_Parent remplies au total")
    return parent_map, parent_map_exact, unmatched


# ── CLI ───────────────────────────────────────────────────────────────
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Injecter les ID parents dans le fichier notes ACS')
    parser.add_argument('--notes',     required=True, help='Fichier Excel des notes')
    parser.add_argument('--schoolify', required=True, help='Fichier Schoolify')
    parser.add_argument('--out',       required=True, help='Fichier de sortie enrichi')
    args = parser.parse_args()

    inject_parent_ids(
        notes_path     = Path(args.notes),
        schoolify_path = Path(args.schoolify),
        out_path       = Path(args.out),
    )
