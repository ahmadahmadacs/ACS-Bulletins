"""
ACS — Core bulletin generator (v7)
Used by both the Streamlit app and CLI.
"""

import os
import copy
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import Color

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
GROUPS = {
    "EB1-2": {"nb_cols": 10},
    "EB3-6": {"nb_cols": 12},
    "EB7":   {"nb_cols": 10},
    "EB8":   {"nb_cols": 11},
    "EB9":   {"nb_cols": 10},
}

SEMESTRES = ["S1", "S2", "S3"]
SEM_ROW   = {"S1": 8, "S2": 9, "S3": 10}

WHITE_COLOR = Color(rgb="FFFFFFFF")
BLACK_COLOR = Color(rgb="FF000000")

# ─────────────────────────────────────────────
# 🖋️  ENSURE CALIBRI
# ─────────────────────────────────────────────
def ensure_calibri(log=print):
    """Vérifie que Carlito/Calibri est disponible (installé via packages.txt)."""
    # Vérification par présence de fichier — pas de subprocess
    font_dirs = [
        "/usr/share/fonts",
        "/usr/local/share/fonts",
        os.path.expanduser("~/.fonts"),
    ]
    found = False
    for d in font_dirs:
        if os.path.isdir(d):
            for root, _, files in os.walk(d):
                if any("carlito" in f.lower() for f in files):
                    found = True
                    break
        if found:
            break

    if found:
        log("✅ Police Carlito (Calibri) disponible")
    else:
        log("⚠️  Carlito non trouvé — vérifiez que packages.txt est présent dans le repo")

    # Alias fontconfig (écriture fichier seulement, pas de subprocess)
    CONF_PATH = "/etc/fonts/conf.d/99-calibri-alias.conf"
    if not os.path.exists(CONF_PATH):
        conf_xml = """\
<?xml version="1.0"?>
<!DOCTYPE fontconfig SYSTEM "fonts.dtd">
<fontconfig>
  <alias>
    <family>Calibri</family>
    <prefer><family>Carlito</family></prefer>
  </alias>
  <alias>
    <family>Calibri Bold</family>
    <prefer><family>Carlito</family></prefer>
  </alias>
</fontconfig>"""
        try:
            with open(CONF_PATH, "w") as f:
                f.write(conf_xml)
            log("✅ Alias fontconfig Calibri→Carlito créé")
        except Exception:
            log("ℹ️  Alias fontconfig non créé (droits insuffisants) — ignoré")
    else:
        log("✅ Alias fontconfig déjà présent")


def ensure_libreoffice(log=print):
    """Vérifie que LibreOffice est disponible (installé via packages.txt)."""
    if shutil.which("libreoffice"):
        log("✅ LibreOffice disponible")
    else:
        log("❌ LibreOffice introuvable — assurez-vous que packages.txt contient 'libreoffice'")

# ─────────────────────────────────────────────
# 🎨 COLOR HELPERS
# ─────────────────────────────────────────────
def copy_color(color_obj):
    if color_obj is None:
        return BLACK_COLOR
    return copy.copy(color_obj)

def is_white_cell(row_idx, col_idx):
    # G5, G6 retirés (v8) — pas de fond propre → texte NOIR
    WHITE_CELLS = {
        (5, 5), (6, 5),    # E5, E6
        (9, 8),            # H9
        (14, 8),           # H14
        (19, 8),           # H19
    }
    for r in (20, 21, 22):
        for c in range(1, 9):
            WHITE_CELLS.add((r, c))
    return (row_idx, col_idx) in WHITE_CELLS

# ─────────────────────────────────────────────
# LECTURE DES NOTES
# ─────────────────────────────────────────────
def load_group_semester(file_path, group, sem):
    sheet_name = f"{sem} Notes {group}"
    wb = load_workbook(file_path, data_only=True, read_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise Exception(f"Sheet manquante : {sheet_name}")

    ws          = wb[sheet_name]
    nb_cols     = GROUPS[group]["nb_cols"]
    result      = {}
    pending     = []
    header_done = False
    current_class_fichier = ""  # nom de classe pour le fichier PDF (col B de معدل الصف)

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        val0 = str(row[0]).strip()

        if val0 == "Nom":
            header_done = True
            continue
        if not header_done:
            if isinstance(row[0], str):
                header_done = True
            else:
                continue

        if isinstance(row[0], (int, float)):
            continue

        if "معدل" in val0:
            avg = []
            for i in range(nb_cols):
                v = row[2 + i] if (2 + i) < len(row) else None
                avg.append(round(v, 2) if isinstance(v, float) else v)

            # Lire le nom de classe en colonne B
            # Détecte معدل الصف avec ou sans variantes d'encodage
            is_classe_row = "الصف" in val0 or val0.strip().startswith("معدل")
            if is_classe_row:
                cn = ""
                # Chercher dans col B (index 1) et col C (index 2) au cas où
                for col_idx in (1, 2):
                    raw = row[col_idx] if col_idx < len(row) else None
                    if raw is not None and str(raw).strip():
                        candidate = str(raw).strip()
                        # Ignorer si c'est un nombre (c'est une note, pas un nom de classe)
                        try:
                            float(candidate)
                        except ValueError:
                            cn = candidate
                            break
                if cn:
                    current_class_fichier = cn
                    for nom in pending:
                        result[nom]["classe_fichier"] = current_class_fichier

            for nom in pending:
                result[nom]["class_avg"] = avg
            pending = []
            continue

        nom    = row[0]
        classe = row[1] if len(row) > 1 else ""  # valeur originale pour le bulletin
        notes  = [row[2 + i] if (2 + i) < len(row) else None for i in range(nb_cols)]
        result[nom] = {"classe": classe, "classe_fichier": current_class_fichier, "notes": notes, "class_avg": None}
        pending.append(nom)

    wb.close()
    return result


def build_students(notes_path, group):
    students = {}
    errors = []
    for sem in SEMESTRES:
        try:
            data = load_group_semester(notes_path, group, sem)
            for nom, info in data.items():
                if nom not in students:
                    students[nom] = {
                        "classe":         info["classe"],
                        "classe_fichier": info.get("classe_fichier", ""),
                        "sems": {}, "avgs": {}
                    }
                # Mettre à jour classe_fichier si on en trouve un non-vide
                if info.get("classe_fichier"):
                    students[nom]["classe_fichier"] = info["classe_fichier"]
                students[nom]["sems"][sem] = info["notes"]
                if info["class_avg"] is not None:
                    students[nom]["avgs"][sem] = info["class_avg"]
        except Exception as e:
            errors.append(f"{group} {sem}: {e}")
    return students, errors

# ─────────────────────────────────────────────
# FIX FORMATTING
# ─────────────────────────────────────────────
def fix_bulletin_formatting(wb):
    ws = wb["Bulletin"]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            old_font = cell.font
            color = WHITE_COLOR if is_white_cell(cell.row, cell.column) \
                    else copy_color(old_font.color)
            cell.font = Font(name="Calibri", size=12,
                             bold=old_font.bold, color=color)

    # Col G rows 9-21 : labels matières
    for row_idx in range(9, 22):
        cell = ws.cell(row=row_idx, column=7)
        if cell.value is not None:
            old_a   = cell.alignment
            fsize   = 9 if len(str(cell.value)) > 12 else 11
            color   = WHITE_COLOR if row_idx >= 20 else copy_color(cell.font.color)
            cell.font = Font(name="Calibri", size=fsize,
                             bold=cell.font.bold, color=color)
            cell.alignment = Alignment(
                horizontal=old_a.horizontal, vertical="center",
                wrap_text=True, shrink_to_fit=True)

    # G5, G6 : pas de fond propre → texte NOIR (v8 fix)
    for row_idx in (5, 6):
        cell = ws.cell(row=row_idx, column=7)
        if cell.value is not None:
            cell.font = Font(name="Calibri", size=12,
                             bold=cell.font.bold, color=BLACK_COLOR)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", shrink_to_fit=True)

    # E5, E6 blanc
    for addr in ("E5", "E6"):
        cell = ws[addr]
        if cell.value is not None:
            cell.font = Font(name="Calibri", size=12,
                             bold=True, color=WHITE_COLOR)
            cell.alignment = Alignment(
                horizontal="right", vertical="center",
                wrap_text=False, shrink_to_fit=True)

    # Row 8 headers
    for col_idx in range(1, 6):
        cell = ws.cell(row=8, column=col_idx)
        if cell.value is not None:
            cell.font = Font(name="Calibri", size=12,
                             bold=True, color=copy_color(cell.font.color))
            cell.alignment = Alignment(
                horizontal="center", vertical="center",
                wrap_text=False, shrink_to_fit=True)


def fix_etudiant_formatting(wb):
    ws = wb["etudiant"]
    ws["B5"].font      = Font(name="Calibri", size=12, bold=True)
    ws["B5"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B4"].font      = Font(name="Calibri", size=12, bold=False)
    ws["B4"].alignment = Alignment(horizontal="right", vertical="center")
    for row_idx in range(7, 11):
        for col_idx in range(3, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.font      = Font(name="Calibri", size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")

# ─────────────────────────────────────────────
# FILL + EXPORT
# ─────────────────────────────────────────────
# Semestres à remplir selon le semestre cible
SEMS_TO_FILL = {
    "S1": ["S1"],
    "S2": ["S1", "S2"],
    "S3": ["S1", "S2", "S3"],
}

# Lignes des matières dans la feuille Bulletin (col B = moyenne, C = S1, D = S2, E = S3)
BULLETIN_MATIERE_ROWS = list(range(9, 22))  # lignes 9 à 21
BULLETIN_COL = {"S1": 3, "S2": 4, "S3": 5}  # C=3, D=4, E=5
BULLETIN_AVG_COL = 2  # col B

def fill_template(template_path, out_path, nom, data, group, semestre_cible="S3"):
    shutil.copy(template_path, out_path)
    wb      = load_workbook(out_path)
    ws_et   = wb["etudiant"]
    ws_bul  = wb["Bulletin"]
    nb_cols = GROUPS[group]["nb_cols"]
    cols    = list(range(3, 3 + nb_cols))

    ws_et["B5"].value = nom
    ws_et["B4"].value = data["classe"]

    sems_actifs = SEMS_TO_FILL.get(semestre_cible, ["S1", "S2", "S3"])

    # ── Feuille etudiant : remplir les notes des semestres actifs ──
    for sem, row_idx in SEM_ROW.items():
        if sem not in sems_actifs:
            continue
        if sem not in data["sems"]:
            continue
        notes = data["sems"][sem]
        for i, col in enumerate(cols):
            if i < len(notes):
                ws_et.cell(row=row_idx, column=col).value = (
                    notes[i] if notes[i] is not None else None)

    # ── Feuille etudiant : moyenne de classe (row 7, lue depuis le fichier notes) ──
    # Copiée pour tous les semestres (S1, S2, S3)
    avg_notes = data["avgs"].get(semestre_cible)
    if avg_notes is None:
        # fallback : semestre actif le plus récent disponible
        for s in reversed(sems_actifs):
            if s in data["avgs"]:
                avg_notes = data["avgs"][s]
                break
    if avg_notes:
        for i, col in enumerate(cols):
            if i < len(avg_notes) and avg_notes[i] is not None:
                ws_et.cell(row=7, column=col).value = avg_notes[i]

    # ── Feuille Bulletin : date d'émission en C6 ──
    ws_bul["C6"].value = date.today().strftime("%d-%m-%Y")

    # ── Feuille Bulletin : moyenne par matière en col B (lignes 9-19) ──
    # Calculée depuis etudiant (colonnes C/D/E = S1/S2/S3)
    # S1 → vide, S2 → (C+D)/2, S3 → (C+D+E)/3  — note vide = 0
    nb_sems = len(sems_actifs)
    MATIERE_ROWS_NORMAL = list(range(9, 20))  # lignes 9 à 19 (depuis etudiant)
    b_values = {}  # stocke les moyennes B9:B19 pour calculer B20 ensuite
    for row_idx in MATIERE_ROWS_NORMAL:
        if semestre_cible == "S1":
            ws_bul.cell(row=row_idx, column=BULLETIN_AVG_COL).value = None
            b_values[row_idx] = 0.0
        else:
            total = 0.0
            for sem in sems_actifs:
                sem_row = SEM_ROW[sem]
                mat_idx = row_idx - 9
                et_col  = cols[mat_idx] if mat_idx < len(cols) else None
                if et_col is not None:
                    v = ws_et.cell(row=sem_row, column=et_col).value
                    total += float(v) if v is not None else 0.0
            moyenne = round(total / nb_sems, 2)
            ws_bul.cell(row=row_idx, column=BULLETIN_AVG_COL).value = moyenne
            b_values[row_idx] = moyenne

    # ── Feuille Bulletin : B20 = somme de B9:B19 ──
    # On utilise b_values déjà calculées, pas ws_bul.cell (qui peut contenir des formules)
    if semestre_cible == "S1":
        ws_bul.cell(row=20, column=BULLETIN_AVG_COL).value = None
    else:
        total_b20 = sum(b_values.values())
        ws_bul.cell(row=20, column=BULLETIN_AVG_COL).value = round(total_b20, 2)

    # B21 : laissé tel quel (formule du template)

    # ── Étendre la zone d'impression jusqu'à la ligne 44 ──
    ws_bul.print_area = "A1:H44"

    fix_bulletin_formatting(wb)
    fix_etudiant_formatting(wb)
    wb.save(out_path)
    wb.close()


def export_pdf(xlsx_path, pdf_path):
    wb = load_workbook(xlsx_path)
    for sheetname in wb.sheetnames:
        wb[sheetname].sheet_state = (
            "visible" if sheetname == "Bulletin" else "hidden")
    wb.save(xlsx_path)
    wb.close()

    outdir = str(Path(pdf_path).parent)
    subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "pdf",
         "--outdir", outdir, xlsx_path],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    generated = Path(outdir) / (Path(xlsx_path).stem + ".pdf")
    if generated.exists():
        generated.rename(pdf_path)
        return True
    return False

# ─────────────────────────────────────────────
# MAIN GENERATION FUNCTION
# ─────────────────────────────────────────────
def generate_all(notes_path, templates: dict, output_dir: Path,
                 log=print, progress_cb=None, semestre_cible="S3"):
    """
    notes_path   : path to the notes Excel file
    templates    : dict {group: path_to_template_xlsx}
    output_dir   : where to write PDFs
    log          : callable for log messages
    progress_cb  : callable(current, total, student_name)
    Returns      : (pdf_list, error_list)
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    tmp_dir  = output_dir / "tmp"
    xlsx_dir = output_dir / "xlsx"
    tmp_dir.mkdir(exist_ok=True)
    xlsx_dir.mkdir(exist_ok=True)

    pdf_list   = []
    xlsx_list  = []
    error_list = []
    all_students = []

    # Collect all students first to know total
    for group, tmpl_path in templates.items():
        students, errs = build_students(notes_path, group)
        error_list.extend(errs)
        for nom, data in students.items():
            all_students.append((group, nom, data, tmpl_path))

    total = len(all_students)
    log(f"📋 {total} bulletins à générer")

    for idx, (group, nom, data, tmpl_path) in enumerate(all_students):
        safe_nom    = nom.replace(" ", "_")
        safe_classe = str(data.get("classe_fichier", "")).replace(" ", "_").replace("/", "-")
        prefix      = safe_classe if safe_classe else group
        xlsx = tmp_dir    / f"{prefix}_{safe_nom}.xlsx"
        pdf  = output_dir / f"{prefix}_{safe_nom}.pdf"

        try:
            fill_template(tmpl_path, xlsx, nom, data, group, semestre_cible=semestre_cible)
            # Copier le xlsx avant conversion (LibreOffice peut modifier le fichier)
            xlsx_out = xlsx_dir / xlsx.name
            shutil.copy2(str(xlsx), str(xlsx_out))
            xlsx_list.append(xlsx_out)
            ok = export_pdf(str(xlsx), str(pdf))
            if ok:
                pdf_list.append(pdf)
                log(f"  ✔ [{idx+1}/{total}] {prefix} — {nom}")
            else:
                error_list.append(f"{nom}: PDF non généré par LibreOffice")
                log(f"  ⚠ [{idx+1}/{total}] {nom} — PDF manquant")
        except Exception as e:
            error_list.append(f"{nom}: {e}")
            log(f"  ❌ [{idx+1}/{total}] {nom} — {e}")

        if progress_cb:
            progress_cb(idx + 1, total, nom)

    shutil.rmtree(tmp_dir, ignore_errors=True)
    return pdf_list, xlsx_list, error_list


def zip_pdfs(pdf_list: list, zip_path: Path):
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for pdf in pdf_list:
            zf.write(pdf, pdf.name)
    return zip_path


def zip_xlsx(xlsx_list: list, zip_path: Path):
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for xlsx in xlsx_list:
            zf.write(xlsx, xlsx.name)
    return zip_path
